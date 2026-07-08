"""Report service - builds tabular datasets and exports to Excel/PDF/CSV."""
import csv
import io
from datetime import date, timedelta
from decimal import Decimal
from typing import Literal

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.money import q_money, to_decimal
from app.models.catalog import Category
from app.models.movement import StockMovement
from app.models.product import Batch, Product
from app.models.supplier import Supplier
from app.repositories.movement import MovementRepository

ExportFormat = Literal["xlsx", "pdf", "csv"]

# --- Report registry: name -> human title ---
REPORTS = {
    "stock_by_product": "Saldo por Produto",
    "below_minimum": "Produtos Abaixo do Estoque Minimo",
    "no_movement": "Produtos sem Movimentacao",
    "near_expiry": "Produtos Proximos do Vencimento",
    "expired": "Produtos Vencidos",
    "movements": "Movimentacoes",
    "suppliers": "Fornecedores",
    "stock_value": "Valor Financeiro em Estoque",
    "stock_by_category": "Saldo por Categoria",
}


class ReportDataset:
    def __init__(self, title: str, headers: list[str], rows: list[list[object]]) -> None:
        self.title = title
        self.headers = headers
        self.rows = rows


class ReportService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.movements = MovementRepository(db)

    # --- dataset builders ---
    def build(self, report: str) -> ReportDataset:
        builder = getattr(self, f"_r_{report}", None)
        if builder is None:
            raise ValueError(f"Relatorio desconhecido: {report}")
        return builder()

    def _all_products(self) -> list[Product]:
        return list(self.db.execute(select(Product)).scalars().all())

    def _r_stock_by_product(self) -> ReportDataset:
        balances = self.movements.all_balances()
        rows = [
            [p.internal_code, p.name, p.unit, float(balances.get(p.id, Decimal("0"))),
             float(p.min_stock or 0), float(p.max_stock or 0)]
            for p in self._all_products()
        ]
        return ReportDataset(
            REPORTS["stock_by_product"],
            ["Codigo", "Produto", "Unidade", "Saldo", "Min", "Max"],
            rows,
        )

    def _r_below_minimum(self) -> ReportDataset:
        balances = self.movements.all_balances()
        rows = [
            [p.internal_code, p.name, float(balances.get(p.id, Decimal("0"))),
             float(p.min_stock or 0)]
            for p in self._all_products()
            if balances.get(p.id, Decimal("0")) < to_decimal(p.min_stock)
        ]
        return ReportDataset(
            REPORTS["below_minimum"], ["Codigo", "Produto", "Saldo", "Minimo"], rows
        )

    def _r_no_movement(self) -> ReportDataset:
        balances = self.movements.all_balances()
        moved = set(balances)
        rows = [
            [p.internal_code, p.name, p.unit]
            for p in self._all_products()
            if p.id not in moved
        ]
        return ReportDataset(REPORTS["no_movement"], ["Codigo", "Produto", "Unidade"], rows)

    def _r_near_expiry(self) -> ReportDataset:
        limit = date.today() + timedelta(days=settings.EXPIRY_ALERT_DAYS)
        stmt = (
            select(Product.internal_code, Product.name, Batch.lot_number, Batch.expiry_date)
            .join(Batch, Batch.product_id == Product.id)
            .where(Batch.expiry_date.isnot(None), Batch.expiry_date <= limit,
                   Batch.expiry_date >= date.today())
            .order_by(Batch.expiry_date)
        )
        rows = [list(r) for r in self.db.execute(stmt).all()]
        return ReportDataset(
            REPORTS["near_expiry"], ["Codigo", "Produto", "Lote", "Validade"], rows
        )

    def _r_expired(self) -> ReportDataset:
        stmt = (
            select(Product.internal_code, Product.name, Batch.lot_number, Batch.expiry_date)
            .join(Batch, Batch.product_id == Product.id)
            .where(Batch.expiry_date.isnot(None), Batch.expiry_date < date.today())
            .order_by(Batch.expiry_date)
        )
        rows = [list(r) for r in self.db.execute(stmt).all()]
        return ReportDataset(REPORTS["expired"], ["Codigo", "Produto", "Lote", "Validade"], rows)

    def _r_movements(self) -> ReportDataset:
        stmt = select(StockMovement).order_by(StockMovement.moved_at.desc()).limit(5000)
        rows = [
            [m.id, m.product_id, m.movement_type.value, m.direction.value,
             float(m.quantity), m.moved_at.strftime("%Y-%m-%d %H:%M"),
             "Sim" if m.is_cancelled else "Nao"]
            for m in self.db.execute(stmt).scalars().all()
        ]
        return ReportDataset(
            REPORTS["movements"],
            ["ID", "Produto", "Tipo", "Direcao", "Qtd", "Data", "Cancelada"],
            rows,
        )

    def _r_suppliers(self) -> ReportDataset:
        rows = [
            [s.cnpj, s.legal_name, s.trade_name or "", s.city or "", s.state or "", s.phone or ""]
            for s in self.db.execute(select(Supplier)).scalars().all()
        ]
        return ReportDataset(
            REPORTS["suppliers"],
            ["CNPJ", "Razao Social", "Nome Fantasia", "Cidade", "UF", "Telefone"],
            rows,
        )

    def _r_stock_value(self) -> ReportDataset:
        """Financial value using the maintained weighted-average cost."""
        balances = self.movements.all_balances()
        rows = []
        total = Decimal("0")
        for p in self._all_products():
            qty = balances.get(p.id, Decimal("0"))
            cost = to_decimal(p.average_cost)
            value = q_money(qty * cost)
            total += value
            rows.append([p.internal_code, p.name, float(qty), float(cost), float(value)])
        rows.append(["", "TOTAL", "", "", float(q_money(total))])
        return ReportDataset(
            REPORTS["stock_value"],
            ["Codigo", "Produto", "Saldo", "Custo Medio", "Valor Total"],
            rows,
        )

    def _r_stock_by_category(self) -> ReportDataset:
        balances = self.movements.all_balances()
        rows_raw = self.db.execute(
            select(Category.name, Product.id).join(Product, Product.category_id == Category.id)
        ).all()
        agg: dict[str, Decimal] = {}
        for name, pid in rows_raw:
            agg[name] = agg.get(name, Decimal("0")) + balances.get(pid, Decimal("0"))
        rows = [[k, float(v)] for k, v in sorted(agg.items())]
        return ReportDataset(REPORTS["stock_by_category"], ["Categoria", "Saldo"], rows)

    # --- exporters ---
    def export(self, report: str, fmt: ExportFormat) -> tuple[bytes, str, str]:
        """Return (content, media_type, filename)."""
        dataset = self.build(report)
        if fmt == "csv":
            return self._to_csv(dataset), "text/csv", f"{report}.csv"
        if fmt == "xlsx":
            return (
                self._to_xlsx(dataset),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"{report}.xlsx",
            )
        if fmt == "pdf":
            return self._to_pdf(dataset), "application/pdf", f"{report}.pdf"
        raise ValueError("Formato invalido")

    @staticmethod
    def _to_csv(ds: ReportDataset) -> bytes:
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=";")
        writer.writerow(ds.headers)
        writer.writerows(ds.rows)
        return buf.getvalue().encode("utf-8-sig")

    @staticmethod
    def _to_xlsx(ds: ReportDataset) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = ds.title[:31]
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for col, header in enumerate(ds.headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for r, row in enumerate(ds.rows, start=2):
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=value)
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(width + 2, 50)
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    @staticmethod
    def _to_pdf(ds: ReportDataset) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

        out = io.BytesIO()
        doc = SimpleDocTemplate(
            out, pagesize=landscape(A4), topMargin=1 * cm, bottomMargin=1 * cm
        )
        data = [ds.headers] + [[str(c) if c is not None else "" for c in row] for row in ds.rows]
        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#EEF3FA")],
                    ),
                ]
            )
        )
        doc.build([table])
        return out.getvalue()
